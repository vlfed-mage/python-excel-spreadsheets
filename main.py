import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    workbook = xl.load_workbook(filename)  # get Excel file
    sheet = workbook['Sheet1']  # get sheet from transactions.xlsx

    cell = sheet['a1']
    print(cell)  # <Cell 'Sheet1'.A1>

    # the same as
    cell = sheet.cell(1, 1)
    print(cell)  # <Cell 'Sheet1'.A1>
    print(cell.value)  # transaction_id

    print(sheet.max_row)  # 4

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9  # calc new value
        corrected_price_cell = sheet.cell(row, 4)  # get new cell in 4th column
        corrected_price_cell.value = corrected_price  # set value to this cell

    values = Reference(  # collect values
        sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4
    )
    chart = BarChart()  # create chart
    chart.add_data(values)  # added data to the chart
    sheet.add_chart(chart, 'e2')  # added chart to the sheet

    workbook.save(filename)  # save changes into file


process_workbook('transactions.xlsx')
